"""
Export utilities для экспорта данных в CSV и Excel
"""

import io
import csv
from typing import List, Dict, Any, Optional
from datetime import date, datetime

from fastapi import Response
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def data_to_csv(data: List[Dict[str, Any]], filename: str = "export.csv") -> StreamingResponse:
    """
    Экспорт данных в CSV
    
    Args:
        data: Список словарей с данными
        filename: Имя файла
    
    Returns:
        StreamingResponse с CSV файлом
    """
    if not data:
        raise ValueError("No data to export")
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    
    writer.writeheader()
    writer.writerows(data)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def data_to_excel(
    data: List[Dict[str, Any]],
    filename: str = "export.xlsx",
    sheet_name: str = "Data",
    column_headers: Optional[Dict[str, str]] = None,
) -> StreamingResponse:
    """
    Экспорт данных в Excel
    
    Args:
        data: Список словарей с данными
        filename: Имя файла
        sheet_name: Название листа
        column_headers: Маппинг заголовков колонок {field: display_name}
    
    Returns:
        StreamingResponse с Excel файлом
    """
    if not data:
        raise ValueError("No data to export")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Заголовки
    headers = list(data[0].keys())
    display_headers = [column_headers.get(h, h) if column_headers else h for h in headers]
    
    # Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавляем заголовки
    ws.append(display_headers)
    
    # Стилізуем заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Добавляем данные
    for row in data:
        values = [row.get(h) for h in headers]
        # Обработка дат и datetime
        processed_values = []
        for v in values:
            if isinstance(v, date):
                processed_values.append(v.strftime('%Y-%m-%d'))
            elif isinstance(v, datetime):
                processed_values.append(v.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                processed_values.append(v)
        ws.append(processed_values)
    
    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# Примеры использования в API
def create_export_router(prefix: str, tags: List[str], get_data_func, entity_name: str):
    """
    Factory для создания роутеров экспорта
    
    Args:
        prefix: Префикс роутера (например, "/employees")
        tags: Теги для OpenAPI
        get_data_func: Функция для получения данных
        entity_name: Название сущности для имен файлов
    """
    from fastapi import APIRouter, Depends, Query
    
    router = APIRouter(prefix=prefix, tags=tags)
    
    @router.get("/export/csv")
    async def export_csv(
        limit: int = Query(1000, le=10000),
        db=Depends(lambda: None),  # Заменить на реальную зависимость
    ):
        """Экспорт в CSV"""
        data = await get_data_func(limit=limit)
        return data_to_csv(data, filename=f"{entity_name}_export_{date.today()}.csv")
    
    @router.get("/export/excel")
    async def export_excel(
        limit: int = Query(1000, le=10000),
        db=Depends(lambda: None),  # Заменить на реальную зависимость
    ):
        """Экспорт в Excel"""
        data = await get_data_func(limit=limit)
        return data_to_excel(
            data,
            filename=f"{entity_name}_export_{date.today()}.xlsx",
            column_headers={
                "id": "ID",
                "full_name": "ФИО",
                "email": "Email",
                "phone": "Телефон",
                "hire_date": "Дата найма",
                "department_id": "Отдел",
                "position_id": "Должность",
            }
        )
    
    return router
